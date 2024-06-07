import openpyxl
import pyodbc 
import sys
# Give the location of the file
filename = str(sys.argv[1])
path = "c:\\python\\HCM2\\input\\" + filename 
 
# To open the workbook 
# workbook object is created
wb_obj = openpyxl.load_workbook(path)
 
# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj["Student Data"]

# Connect to the database using a DSN
cnxn = pyodbc.connect("DSN=nsad;UID=applogin;PWD=~!@#$Jessam00")
sql = "truncate table cns_stagging.dbo.HCM2_Student"
cursor = cnxn.cursor()
cursor.execute(sql)
cursor.commit()
maxrow = sheet_obj.max_row
awardyear = sheet_obj.cell(row=4,column=2).value.strip().replace("-20","-")

print("maxrow = ",maxrow)
print("awardyear = ",awardyear)



for i in range(8,maxrow + 1):

    sequence_number = sheet_obj.cell(row=i,column=1).value.strip()
    last_name = sheet_obj.cell(row=i,column=2).value.strip().replace("'","''")
    first_name = sheet_obj.cell(row=i,column=3).value.strip().replace("'","''")
    ssn = sheet_obj.cell(row=i,column=4).value.strip()
    address = sheet_obj.cell(row=i,column=5).value.strip()
    telephone = sheet_obj.cell(row=i,column=6).value.strip()


    sql = "insert into cns_stagging.dbo.HCM2_Student (seq_nbr, last_name, first_name, ssn, telephone, awardyear, filename) values (" + str(sequence_number) + ", '" + last_name + "','" + first_name + "','" + ssn + "','" + telephone + "','" + awardyear + "','" + filename + "' )"
    cursor.execute(sql)
    cursor.commit()


